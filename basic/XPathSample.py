#! python
# XPathSample.py
# 入力内容をバッファリングし、XPathで指定した値を取得する
# 以下がとても参考になった
# https://vaaaaaanquish.hatenablog.com/entry/2017/06/25/202924
#
import requests     # pip install requests
import bs4          # pip install beautifulsoup4
import sys          # nessesaly nead bs4
import os           # nessesaly nead bs4
import logging
from logging import DEBUG,INFO,WARN,ERROR,CRITICAL
from lxml import html    # pip install lxml
import json
import copy
import openpyxl     # pip install openpyxl


# ロギング初期化
logging.basicConfig(
        level=logging.DEBUG,            # ログレベル
        format=' %(asctime)s - %(levelname)s - %(lineno)s - %(message)s')

# 銘柄情報見出し部分
MEIGARA_HEAD = "//table[@class='tbl_dataOutputloop_02']/tr[1]/*"
# 銘柄情報データ部分
MEIGARA_BODY = "//table[@class='tbl_dataOutputloop_02']/tr[2]"

# 銘柄一覧見出し
meigara_head = ['取得日']

# --------------------------------------------------------
# json内容をファイル出力
# --------------------------------------------------------
def filewrite(meigara_dic_list):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '銘柄一覧'

    # 初期位置A1から右に見出しを追記
    work_r = 1
    work_c = 1
    logging.debug("len : {}".format(len(meigara_head)))
    for i in range(len(meigara_head)):
        sheet.cell(row=work_r, column=work_c).value = meigara_head[i]
        work_c += 1

    # 見出しに沿ってデータを追記
    # 初期位置A2
    work_r = 2
    work_c = 1
    logging.debug("meigara_dic_list type:{} data:{}".format(type(meigara_dic_list), meigara_dic_list))
    for dic_item in meigara_dic_list:
        logging.debug("dic_item : {}".format(dic_item))
        work_c = 1
        for i in range(len(meigara_head)):
            # 1列名の取得日はデータを取得した内容ではなく、本日の日付を設定する
            if( i == 0 ):
                sheet.cell(row=work_r, column=work_c).value = 'yyyy/mm/dd'
            else:
                # 列名に対応する値を取得し、セルに設定
                logging.debug("i : {}".format(i))
                logging.debug("meigara_head[i] : {}".format(meigara_head[i]))
                sheet.cell(row=work_r, column=work_c).value = dic_item[meigara_head[i]]
            work_c += 1
        work_r += 1

    wb.save('output.xlsx')

# --------------------------------------------------------
# 汎用Xpath
# --------------------------------------------------------
def getXpathList(dom,xpathStr):
    logging.debug('=== [{}] start ==='.format(sys._getframe().f_code.co_name))
    result = []

    list = dom.xpath(xpathStr)
    logging.debug('== xpath:{} list len:{}'.format(xpathStr, len(list)))
    if list is not None:
        for obj in list:
            logging.debug('== tag:{} text:{} attr:{}'.format(obj.tag, obj.text, obj.attrib))

    return result

# --------------------------------------------------------
# 各銘柄を取得し、辞書リスト化 ver3
# --------------------------------------------------------
def databody3(dom):

    listXpath = "//table[@class='tbl_dataOutputloop_02']//tr"
#    getXpathList(dom, listXpath)
    listbodyxpaths = [
        "/td[1]/a",
        "/td[2]",
        "/td[3]",
        "/td[4]/table/tr/td[@class='align_R']",
        "/td[5]",
        "/td[6]/span[2]",
        "/td[7]"
    ]

    domlist = dom.xpath(listXpath)

    listhead = []
    tbody = {}
    result = []
    tridx = 0
#    for tridx in range(len(domlist)):
    for tridx in range(10):

        headidx = 0
        # 最後のtrタグは評価合計なので集計対象外
        logging.debug('== len:{}'.format(len(domlist)))
        if(tridx == (len(domlist)-1)):
            break

        # 最初のtrタグは見出し
        if(tridx == 0):
            trobj = domlist[tridx]
            for tdobj in trobj:
                logging.debug('== tag:{} text:{}'.format(tdobj.tag, tdobj.text.strip()))
                listhead.append(tdobj.text.strip())
                headidx += 1
            continue

        # 2番目以降のtrタグはテーブルデータ
        # 取得するXpathが異なるので再度domを取り直す
        for headidx in range(len(listhead)):

            logging.debug('== tridx:{} headidx:{}'.format(tridx, headidx))
            bodyxpath = listXpath + '[' + str(tridx+1) + ']' + listbodyxpaths[headidx]
            logging.debug('== xpath:{}'.format(bodyxpath))
            bodylist = dom.xpath(bodyxpath)

            # 1要素目のみ取得
            obj = bodylist[0]
            logging.debug('== tag:{} text:{}'.format(obj.tag, obj.text.strip()))
            tbody[listhead[headidx]] = obj.text.strip()

        # tr1要素分を戻り値リストの1要素として登録
        result.append(copy.deepcopy(tbody))

    for headobj in listhead:
        meigara_head.append(headobj)

    return result


# --------------------------------------------------------
# 各銘柄を取得し、辞書リスト化 ver2
# --------------------------------------------------------
def databody2(dom):
    logging.debug('=== [{}] start ==='.format(sys._getframe().f_code.co_name))
    # 最終的なデータ構成イメージ
    # {
    #    '銘柄':'',
    #    '保有数量':'',
    #    '売却可能数量':'',
    #    '平均取得単価':'',
    #    '取得金額':'',
    #    '評価損益':'',
    # }

    result = {}

    # 見出しと値を取得するXPath
    listhead = dom.xpath(MEIGARA_HEAD)
    listbodyxpaths = [
        MEIGARA_BODY + "/td[1]/a",
        MEIGARA_BODY + "/td[2]",
        MEIGARA_BODY + "/td[3]",
        MEIGARA_BODY + "/td[4]/table/tr/td[@class='align_R']",
        MEIGARA_BODY + "/td[5]",
        MEIGARA_BODY + "/td[6]/span[@class='valueminus']"
    ]

    # XPathで取得した内容をJson形式に割り当てる
    idx = 0
    for idx in range(len(listbodyxpaths)):
        bodylist = dom.xpath(listbodyxpaths[idx])
        for obj in bodylist:
            logging.debug('== tag:{} text:{}'.format(obj.tag, obj.text.strip()))
            result[listhead[idx].text.strip()] = obj.text.strip()

        # 他で使うので、銘柄一覧の見出し文字列は保持する
        meigara_head.append(listhead[idx].text.strip())

        idx += 1

    return result

# --------------------------------------------------------
# 各銘柄を取得し、リスト化
# ver2ができたことにより、未使用
# --------------------------------------------------------
def databody(dom):
    logging.debug('=== [{}] start ==='.format(sys._getframe().f_code.co_name))
    # 最終的なデータ構成イメージ
    # {
    #    '銘柄':
    #     {
    #         '保有数量':'',
    #         '売却可能数量':'',
    #         '平均取得単価':'',
    #         '取得金額':'',
    #         '評価損益':'',
    #     }
    # }

    result = {}

    listhead = dom.xpath(MEIGARA_HEAD)
#    list = dom.xpath(MEIGARA_BODY)
#    for obj in list:
#        logging.debug('== tag:{} text:{}'.format(obj.tag, obj.text.strip()))

    # 銘柄
    list = dom.xpath(MEIGARA_BODY + "/td[1]/a")
    for obj in list:
        logging.debug('== tag:{} text:{}'.format(obj.tag, obj.text.strip()))
        result[listhead[1].text] = obj.text.strip()

    # 保有数量
    list = dom.xpath(MEIGARA_BODY + "/td[2]")
    for obj in list:
        logging.debug('== tag:{} text:{}'.format(obj.tag, obj.text.strip()))
        result[listhead[2]] = obj.text.strip()

    # 売却可能数量
    list = dom.xpath(MEIGARA_BODY + "/td[3]")
    for obj in list:
        logging.debug('== tag:{} text:{}'.format(obj.tag, obj.text.strip()))
        result[listhead[3]] = obj.text.strip()

    # 平均取得単価
    list = dom.xpath(MEIGARA_BODY + "/td[4]/table/tr/td[@class='align_R']")
    for obj in list:
        logging.debug('== tag:{} text:{}'.format(obj.tag, obj.text.strip()))
        result[listhead[4]] = obj.text.strip()

    # 取得金額
    list = dom.xpath(MEIGARA_BODY + "/td[5]")
    for obj in list:
        logging.debug('== tag:{} text:{}'.format(obj.tag, obj.text.strip()))
        result[listhead[5]] = obj.text.strip()

    # 評価損益
    list = dom.xpath(MEIGARA_BODY + "/td[6]/span[@class='valueminus']")
    for obj in list:
        logging.debug('== tag:{} text:{}'.format(obj.tag, obj.text.strip()))
        result[listhead[6]] = obj.text.strip()

    return result

# --------------------------------------------------------
# 練習
# --------------------------------------------------------
def study(dom):
        # ルート要素
        localxpath = '/*'
        get(dom, localxpath)

        # 属性
        localxpath = "//div[@id='smenu_AstAdp']/a"
        get(dom, localxpath)

        # 特定の要素（タイトル）
        logging.debug('== タイトル取得')
        #list = dom.xpath('//title')
        list = dom.xpath("//a[@title = 'ログアウト']")
        logging.debug('list idx:{}'.format(len(list)))
        for obj in list:
            logging.debug('== tag:{} text:{}'.format(obj.tag, obj.text))


# --------------------------------------------------------
# 処理本体
# --------------------------------------------------------
def targetRun(filepath):
    logging.debug('=== [{}] start ==='.format(sys._getframe().f_code.co_name))

    file1 = open(filepath, mode='r', encoding='shift_jis')
    file2 = open(filepath, mode='r', encoding='shift_jis')
    with file1, file2:
        # ---------------------
        # domオブジェクトの生成
        # ---------------------
        soup = bs4.BeautifulSoup(file1, features='html.parser')
        # fromstringを使用する場合はstr型ではなくエンコードしたbyte型で渡す必要がある
        text = file2.read().encode('shift_jis')
        #logging.info('file2 type = {}'.format(type(text)))
        dom = html.fromstring(text)
        # parseを使用する場合はファイルパスを渡す
        #dom = html.parse(filepath)
        #logging.info('type = {} dom = {} '.format(type(dom), dom))

        # Xpath確認
        getXpathList(dom, "//table[@class='tbl_dataOutputloop_02']//tr[1]/td[1]/a")
        # -----------------------------
        # domオブジェクトを使用した処理
        # -----------------------------
        meigara_dic = databody3(dom)
        logging.info('== databody3 result:{}'.format(meigara_dic))

        # 辞書型のみ対応
        # ensure_ascii=Falseでunicodeエスケープをしない
        #logging.info('== result:{}'.format(json.dumps(meigara_dic, ensure_ascii=False)))


    return meigara_dic

# --------------------------------------------------------
# main
# --------------------------------------------------------
# テストデータパス
targetdatapath = './TestData/sample.html'

# テストメソッド呼び出し
logging.debug(os.getcwd())
meigara = targetRun(targetdatapath)
logging.info('== targetrun result:{}'.format(meigara))

# 引数はリスト型に辞書を入れた形式
filewrite(meigara)



